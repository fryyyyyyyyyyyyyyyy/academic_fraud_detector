"""XLSX 原始数据加载测试。"""

from openpyxl import Workbook

from academic_fraud_detector.utils.raw_data_loader import load_raw_data_files


def test_load_raw_data_files_extracts_column_and_decimal_metadata(tmp_path):
    xlsx_path = tmp_path / "Source Data Fig.1.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Fig1"
    ws["A1"] = "Sample"
    ws["B1"] = "Treatment response"
    for row, value in enumerate([1.25, 2.35, 3.45, 4.55], start=2):
        ws.cell(row=row, column=1).value = f"S{row - 1}"
        cell = ws.cell(row=row, column=2)
        cell.value = value
        cell.number_format = "0.00"
    wb.save(xlsx_path)

    payload = load_raw_data_files([str(xlsx_path)])

    assert payload["profile"]["loaded_file_count"] == 1
    assert payload["profile"]["sheet_count"] == 1
    assert payload["profile"]["dataset_count"] >= 1

    column_dataset = next(
        dataset for dataset in payload["datasets"]
        if dataset["source"]["orientation"] == "column" and dataset["source"]["range"] == "B2:B5"
    )
    assert column_dataset["values"] == [1.25, 2.35, 3.45, 4.55]
    assert column_dataset["raw_values"] == ["1.25", "2.35", "3.45", "4.55"]
    assert column_dataset["decimal_places"] == [2, 2, 2, 2]
    assert column_dataset["last_decimal_digits"] == [5, 5, 5, 5]
    assert column_dataset["decimal_suffixes"]["2"] == ["25", "35", "45", "55"]
    assert column_dataset["row_labels"] == ["S1", "S2", "S3", "S4"]


def test_load_raw_data_files_uses_excel_percent_display_for_decimal_metadata(tmp_path):
    xlsx_path = tmp_path / "percent.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Percent"
    ws["A1"] = "Sample"
    ws["B1"] = "Rate"
    for row, value in enumerate([0.123, 0.456, 0.789], start=2):
        ws.cell(row=row, column=1).value = f"S{row - 1}"
        cell = ws.cell(row=row, column=2)
        cell.value = value
        cell.number_format = "0.0%"
    wb.save(xlsx_path)

    payload = load_raw_data_files([str(xlsx_path)])
    column_dataset = next(
        dataset for dataset in payload["datasets"]
        if dataset["source"]["orientation"] == "column" and dataset["source"]["range"] == "B2:B4"
    )

    assert column_dataset["raw_values"] == ["12.3%", "45.6%", "78.9%"]
    assert column_dataset["last_decimal_digits"] == [3, 6, 9]


def test_load_raw_data_files_counts_gaps_inside_reported_range(tmp_path):
    xlsx_path = tmp_path / "gaps.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Gaps"
    ws["B1"] = "Signal"
    ws["B2"] = 1.1
    ws["B3"] = "N/A"
    ws["B4"] = 2.2
    ws["B6"] = 3.3
    wb.save(xlsx_path)

    payload = load_raw_data_files([str(xlsx_path)])
    column_dataset = next(
        dataset for dataset in payload["datasets"]
        if dataset["source"]["orientation"] == "column" and dataset["source"]["range"] == "B2:B6"
    )

    assert column_dataset["missing_count"] == 2
    assert column_dataset["source"]["analyzed_cells"] == ["B2", "B4", "B6"]


def test_load_raw_data_files_splits_spatially_separated_column_tables(tmp_path):
    xlsx_path = tmp_path / "column_tables.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Mixed"
    ws["A1"] = "Fig.1a"
    for row, value in enumerate([1.1, 1.2, 1.3, 1.4], start=2):
        ws.cell(row=row, column=2).value = value
    ws["A11"] = "Fig.1b"
    for row, value in enumerate([2.1, 2.2, 2.3, 2.4], start=12):
        ws.cell(row=row, column=2).value = value
    wb.save(xlsx_path)

    payload = load_raw_data_files([str(xlsx_path)])
    column_datasets = [
        dataset for dataset in payload["datasets"]
        if dataset["source"]["orientation"] == "column"
    ]

    ranges = {dataset["source"]["range"] for dataset in column_datasets}
    assert "B2:B5" in ranges
    assert "B12:B15" in ranges
    assert "B2:B15" not in ranges

    first = next(dataset for dataset in column_datasets if dataset["source"]["range"] == "B2:B5")
    second = next(dataset for dataset in column_datasets if dataset["source"]["range"] == "B12:B15")
    assert first["source"]["table_title"] == "Fig.1a"
    assert second["source"]["table_title"] == "Fig.1b"


def test_load_raw_data_files_splits_column_tables_on_nearby_title_with_small_gap(tmp_path):
    xlsx_path = tmp_path / "nearby_title.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Fig1"
    ws["H1"] = "Fig.1g"
    for row, value in enumerate([1, 2, 3, 4], start=2):
        ws.cell(row=row, column=8).value = value
    ws["H7"] = "Fig.1h"
    for row, value in enumerate([5, 6, 7, 8], start=8):
        ws.cell(row=row, column=8).value = value
    wb.save(xlsx_path)

    payload = load_raw_data_files([str(xlsx_path)])
    column_datasets = [
        dataset for dataset in payload["datasets"]
        if dataset["source"]["orientation"] == "column"
    ]

    ranges = {dataset["source"]["range"] for dataset in column_datasets}
    assert "H2:H5" in ranges
    assert "H8:H11" in ranges
    assert "H2:H11" not in ranges


def test_load_raw_data_files_splits_spatially_separated_row_tables(tmp_path):
    xlsx_path = tmp_path / "row_tables.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Mixed"
    ws["B1"] = "Fig.2a"
    for col, value in enumerate([1.1, 1.2, 1.3, 1.4], start=2):
        ws.cell(row=3, column=col).value = value
    ws["I1"] = "Fig.2b"
    for col, value in enumerate([2.1, 2.2, 2.3, 2.4], start=9):
        ws.cell(row=3, column=col).value = value
    wb.save(xlsx_path)

    payload = load_raw_data_files([str(xlsx_path)])
    row_datasets = [
        dataset for dataset in payload["datasets"]
        if dataset["source"]["orientation"] == "row"
    ]

    ranges = {dataset["source"]["range"] for dataset in row_datasets}
    assert "B3:E3" in ranges
    assert "I3:L3" in ranges
    assert "B3:L3" not in ranges

    first = next(dataset for dataset in row_datasets if dataset["source"]["range"] == "B3:E3")
    second = next(dataset for dataset in row_datasets if dataset["source"]["range"] == "I3:L3")
    assert first["source"]["table_title"] == "Fig.2a"
    assert second["source"]["table_title"] == "Fig.2b"


def test_load_raw_data_files_reports_missing_file(tmp_path):
    missing = tmp_path / "missing.xlsx"

    payload = load_raw_data_files([str(missing)])

    assert payload["profile"]["loaded_file_count"] == 0
    assert payload["profile"]["errors"]
    assert payload["files"][0]["status"] == "error"


def test_load_raw_data_files_preserves_fixed_decimal_display(tmp_path):
    xlsx_path = tmp_path / "fixed_decimal.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Fixed"
    ws["A1"] = "Sample"
    ws["B1"] = "Signal"
    for row, value in enumerate([1.2, 2.3, 3.4], start=2):
        ws.cell(row=row, column=1).value = f"S{row - 1}"
        cell = ws.cell(row=row, column=2)
        cell.value = value
        cell.number_format = "0.00"
    wb.save(xlsx_path)

    payload = load_raw_data_files([str(xlsx_path)])
    column_dataset = next(
        dataset for dataset in payload["datasets"]
        if dataset["source"]["orientation"] == "column" and dataset["source"]["range"] == "B2:B4"
    )

    assert column_dataset["raw_values"] == ["1.20", "2.30", "3.40"]
    assert column_dataset["decimal_suffixes"]["2"] == ["20", "30", "40"]


def test_load_raw_data_files_flags_expanded_design_variable_keywords(tmp_path):
    xlsx_path = tmp_path / "design_keywords.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Design"
    ws["A1"] = "Sample"
    ws["B1"] = "Treatment group batch"
    for row, value in enumerate([1, 2, 3], start=2):
        ws.cell(row=row, column=1).value = f"S{row - 1}"
        ws.cell(row=row, column=2).value = value
    wb.save(xlsx_path)

    payload = load_raw_data_files([str(xlsx_path)])
    column_dataset = next(
        dataset for dataset in payload["datasets"]
        if dataset["source"]["orientation"] == "column" and dataset["source"]["range"] == "B2:B4"
    )

    assert column_dataset["is_designed_sequence_candidate"]
