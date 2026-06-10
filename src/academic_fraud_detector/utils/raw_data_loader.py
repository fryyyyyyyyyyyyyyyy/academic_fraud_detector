"""XLSX 原始数据加载与标准化。"""

from __future__ import annotations

import math
import re
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - 由调用方在 profile.errors 中处理
    load_workbook = None  # type: ignore[assignment]
    get_column_letter = None  # type: ignore[assignment]

NUMERIC_RE = re.compile(r"^[+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?%?$")
MAX_COLUMN_DATASET_CANDIDATES_PER_SHEET = 500
MAX_ROW_DATASET_CANDIDATES_PER_SHEET = 500

DESIGNED_KEYWORDS = (
    "id",
    "no",
    "编号",
    "序号",
    "sample no",
    "样本编号",
    "time",
    "时间",
    "day",
    "week",
    "dose",
    "dosage",
    "concentration",
    "浓度",
    "剂量",
    "年龄",
    "age",
)


def _is_missing(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _coerce_number(value: Any) -> float | None:
    """将 Excel 单元格值安全转换为 float，排除日期、布尔和明显非数值文本。"""
    if _is_missing(value) or isinstance(value, bool):
        return None
    if isinstance(value, (datetime, date, time)):
        return None
    if isinstance(value, (int, float)):
        if math.isnan(float(value)) or math.isinf(float(value)):
            return None
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if not NUMERIC_RE.match(text):
            return None
        is_percent = text.endswith("%")
        if is_percent:
            text = text[:-1]
        try:
            number = float(text)
        except ValueError:
            return None
        return number / 100 if is_percent else number
    return None


def _decimals_from_format(number_format: str | None) -> int | None:
    """从 Excel number_format 中推断显示小数位数。"""
    if not number_format or number_format == "General":
        return None
    fmt = number_format.split(";")[0]
    if "." not in fmt:
        return 0
    decimal_part = fmt.split(".", 1)[1]
    decimal_part = re.split(r"[%E_\s]", decimal_part)[0]
    digits = [c for c in decimal_part if c in "0#?"]
    if not digits:
        return 0
    return len(digits)


def _display_text(value: Any, number_format: str | None) -> str:
    """尽量重建 Excel 中用于小数末位分析的显示文本。"""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (datetime, date, time)):
        return str(value)
    if isinstance(value, (int, float)):
        decimals = _decimals_from_format(number_format)
        is_percent = bool(number_format and "%" in number_format)
        display_value = float(value) * 100 if is_percent else float(value)
        if decimals is not None and decimals >= 0:
            rendered = f"{display_value:.{decimals}f}"
        else:
            rendered = format(display_value, ".15g")
        return f"{rendered}%" if is_percent else rendered
    return str(value).strip()


def _decimal_metadata(raw_text: str) -> dict[str, Any]:
    """提取小数位、小数最后一位和后缀。"""
    text = raw_text.strip().replace(",", "")
    if text.endswith("%"):
        text = text[:-1]
    if "e" in text.lower():
        try:
            text = format(float(text), ".15f").rstrip("0").rstrip(".")
        except ValueError:
            pass
    signless = text.lstrip("+-")
    if "." not in signless:
        return {
            "decimal_places": 0,
            "last_decimal_digit": None,
            "decimal_part": "",
            "decimal_suffix_1": None,
            "decimal_suffix_2": None,
            "decimal_suffix_3": None,
        }
    decimal_part = re.sub(r"\D", "", signless.split(".", 1)[1])
    if not decimal_part:
        return {
            "decimal_places": 0,
            "last_decimal_digit": None,
            "decimal_part": "",
            "decimal_suffix_1": None,
            "decimal_suffix_2": None,
            "decimal_suffix_3": None,
        }
    return {
        "decimal_places": len(decimal_part),
        "last_decimal_digit": int(decimal_part[-1]),
        "decimal_part": decimal_part,
        "decimal_suffix_1": decimal_part[-1:],
        "decimal_suffix_2": decimal_part[-2:] if len(decimal_part) >= 2 else None,
        "decimal_suffix_3": decimal_part[-3:] if len(decimal_part) >= 3 else None,
    }


def _text_cell(value: Any) -> str | None:
    if isinstance(value, str) and value.strip():
        return value.strip()
    return None


def _nearby_column_header(sheet: Any, row: int, col: int) -> str:
    """寻找数值列上方最近的文本表头。"""
    for r in range(row - 1, max(0, row - 6), -1):
        text = _text_cell(sheet.cell(r, col).value)
        if text:
            return text
    first_row = _text_cell(sheet.cell(1, col).value)
    return first_row or f"Column {get_column_letter(col)}"


def _nearby_row_header(sheet: Any, row: int, first_col: int) -> str:
    """寻找数值行左侧最近的文本行标签。"""
    for c in range(first_col - 1, max(0, first_col - 6), -1):
        text = _text_cell(sheet.cell(row, c).value)
        if text:
            return text
    first_col_text = _text_cell(sheet.cell(row, 1).value)
    return first_col_text or f"Row {row}"


def _row_label(sheet: Any, row: int, value_col: int) -> str:
    for c in range(value_col - 1, max(0, value_col - 5), -1):
        text = _text_cell(sheet.cell(row, c).value)
        if text:
            return text
    return f"Row {row}"


def _col_label(sheet: Any, value_row: int, col: int) -> str:
    for r in range(value_row - 1, max(0, value_row - 5), -1):
        text = _text_cell(sheet.cell(r, col).value)
        if text:
            return text
    return f"Column {get_column_letter(col)}"


def _is_designed_sequence_label(label: str) -> bool:
    lowered = label.lower()
    return any(keyword in lowered for keyword in DESIGNED_KEYWORDS)


def _cell_record(cell: Any) -> dict[str, Any] | None:
    number = _coerce_number(cell.value)
    if number is None:
        return None
    raw_text = _display_text(cell.value, cell.number_format)
    meta = _decimal_metadata(raw_text)
    return {
        "address": cell.coordinate,
        "row": cell.row,
        "column": cell.column,
        "value": number,
        "raw_value": raw_text,
        "number_format": cell.number_format,
        **meta,
    }


def _build_dataset(
    *,
    file_path: Path,
    sheet_name: str,
    orientation: str,
    label: str,
    records: list[dict[str, Any]],
    row_labels: list[str],
    column_labels: list[str],
    header: str,
) -> dict[str, Any] | None:
    if len(records) < 3:
        return None
    first = records[0]["address"]
    last = records[-1]["address"]
    cell_range = first if first == last else f"{first}:{last}"
    analyzed_cells = [r["address"] for r in records]
    if orientation == "column":
        span_count = records[-1]["row"] - records[0]["row"] + 1
    elif orientation == "row":
        span_count = records[-1]["column"] - records[0]["column"] + 1
    else:
        span_count = len(records)
    missing_count = max(0, span_count - len(records))
    dataset_id = f"{file_path.name}::{sheet_name}::{orientation}::{cell_range}"
    decimal_places = [r["decimal_places"] for r in records]
    last_digits = [r["last_decimal_digit"] for r in records]
    suffixes_1 = [r["decimal_suffix_1"] for r in records]
    suffixes_2 = [r["decimal_suffix_2"] for r in records]
    suffixes_3 = [r["decimal_suffix_3"] for r in records]
    values = [r["value"] for r in records]
    raw_values = [r["raw_value"] for r in records]
    designed = _is_designed_sequence_label(label) or _is_designed_sequence_label(header)
    return {
        "dataset_id": dataset_id,
        "source": {
            "file_path": str(file_path),
            "file_name": file_path.name,
            "sheet": sheet_name,
            "orientation": orientation,
            "range": cell_range,
            "analyzed_cells": analyzed_cells,
            "header": header,
        },
        "label": label,
        "row_labels": row_labels,
        "column_labels": column_labels,
        "values": values,
        "raw_values": raw_values,
        "n": len(values),
        "missing_count": missing_count,
        "decimal_places": decimal_places,
        "last_decimal_digits": last_digits,
        "decimal_suffixes": {
            "1": suffixes_1,
            "2": suffixes_2,
            "3": suffixes_3,
        },
        "cells": [
            {
                "address": r["address"],
                "raw_value": r["raw_value"],
                "value": r["value"],
                "number_format": r["number_format"],
            }
            for r in records
        ],
        "is_designed_sequence_candidate": designed,
    }


def _iter_column_records(sheet: Any, col: int) -> Iterable[dict[str, Any]]:
    for row in range(1, sheet.max_row + 1):
        record = _cell_record(sheet.cell(row, col))
        if record is not None:
            yield record


def _iter_row_records(sheet: Any, row: int) -> Iterable[dict[str, Any]]:
    for col in range(1, sheet.max_column + 1):
        record = _cell_record(sheet.cell(row, col))
        if record is not None:
            yield record


def _datasets_from_sheet(file_path: Path, sheet: Any, min_values: int) -> list[dict[str, Any]]:
    datasets: list[dict[str, Any]] = []
    sheet_name = sheet.title
    numeric_by_col: dict[int, list[dict[str, Any]]] = {}
    numeric_by_row: dict[int, list[dict[str, Any]]] = {}
    text_by_cell: dict[tuple[int, int], str] = {}

    # One pass over the worksheet is much faster than repeatedly calling
    # sheet.cell() across max_row × max_column, especially for source-data files
    # with large formatted but mostly empty ranges.
    for row_cells in sheet.iter_rows():
        for cell in row_cells:
            row = getattr(cell, "row", None)
            col = getattr(cell, "column", None)
            if row is None or col is None:
                continue
            text = _text_cell(cell.value)
            if text:
                text_by_cell[(row, col)] = text
            record = _cell_record(cell)
            if record is None:
                continue
            if col in numeric_by_col or len(numeric_by_col) < MAX_COLUMN_DATASET_CANDIDATES_PER_SHEET:
                numeric_by_col.setdefault(col, []).append(record)
            if row in numeric_by_row or len(numeric_by_row) < MAX_ROW_DATASET_CANDIDATES_PER_SHEET:
                numeric_by_row.setdefault(row, []).append(record)

    def nearby_column_header(row: int, col: int) -> str:
        for header_row in range(row - 1, max(0, row - 6), -1):
            text = text_by_cell.get((header_row, col))
            if text:
                return text
        return text_by_cell.get((1, col), f"Column {get_column_letter(col)}")

    def nearby_row_header(row: int, first_col: int) -> str:
        for header_col in range(first_col - 1, max(0, first_col - 6), -1):
            text = text_by_cell.get((row, header_col))
            if text:
                return text
        return text_by_cell.get((row, 1), f"Row {row}")

    def row_label(row: int, value_col: int) -> str:
        for label_col in range(value_col - 1, max(0, value_col - 5), -1):
            text = text_by_cell.get((row, label_col))
            if text:
                return text
        return f"Row {row}"

    def col_label(value_row: int, col: int) -> str:
        for label_row in range(value_row - 1, max(0, value_row - 5), -1):
            text = text_by_cell.get((label_row, col))
            if text:
                return text
        return f"Column {get_column_letter(col)}"

    for col, records in sorted(numeric_by_col.items()):
        if len(records) < min_values:
            continue
        first_row = records[0]["row"]
        header = nearby_column_header(first_row, col)
        col_letter = get_column_letter(col)
        label = f"{file_path.name}/{sheet_name}/{col_letter} {header}"
        row_labels = [row_label(r["row"], col) for r in records]
        dataset = _build_dataset(
            file_path=file_path,
            sheet_name=sheet_name,
            orientation="column",
            label=label,
            records=records,
            row_labels=row_labels,
            column_labels=[header],
            header=header,
        )
        if dataset is not None:
            datasets.append(dataset)

    for row, records in sorted(numeric_by_row.items()):
        if len(records) < min_values:
            continue
        first_col = records[0]["column"]
        header = nearby_row_header(row, first_col)
        label = f"{file_path.name}/{sheet_name}/Row {row} {header}"
        column_labels = [col_label(row, r["column"]) for r in records]
        dataset = _build_dataset(
            file_path=file_path,
            sheet_name=sheet_name,
            orientation="row",
            label=label,
            records=records,
            row_labels=[header],
            column_labels=column_labels,
            header=header,
        )
        if dataset is not None:
            datasets.append(dataset)

    return datasets


def load_raw_data_files(
    raw_data_paths: list[str],
    min_values: int = 3,
    max_datasets_per_file: int = 1000,
    max_total_datasets: int = 10000,
) -> dict[str, Any]:
    """读取本地 XLSX 原始数据文件并返回标准化 dataset payload。

    Source Data 工作簿可能包含上百万数值单元格。为避免把大模型上下文和两两比对
    拖垮，默认按文件和总量截断标准化 dataset，并在 profile.warnings 中显式记录。
    """
    payload: dict[str, Any] = {
        "files": [],
        "datasets": [],
        "profile": {
            "file_count": len(raw_data_paths),
            "loaded_file_count": 0,
            "sheet_count": 0,
            "dataset_count": 0,
            "numeric_value_count": 0,
            "errors": [],
            "warnings": [],
        },
    }

    if load_workbook is None:
        error = "openpyxl 未安装，无法解析 XLSX 原始数据。"
        payload["profile"]["errors"].append(error)
        return payload

    for raw_path in raw_data_paths:
        path = Path(raw_path).expanduser()
        file_entry: dict[str, Any] = {
            "path": str(path),
            "file_name": path.name,
            "status": "pending",
            "sheet_names": [],
            "dataset_count": 0,
            "errors": [],
        }
        payload["files"].append(file_entry)

        if len(payload["datasets"]) >= max_total_datasets:
            file_entry["status"] = "skipped"
            warning = f"已达到全局 dataset 上限 {max_total_datasets}；未读取 {path.name}。"
            file_entry.setdefault("warnings", []).append(warning)
            payload["profile"]["warnings"].append(warning)
            continue

        if not path.exists():
            file_entry["status"] = "error"
            file_entry["errors"].append(f"原始数据文件不存在: {path}")
            payload["profile"]["errors"].append(file_entry["errors"][-1])
            continue
        if path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            file_entry["status"] = "error"
            file_entry["errors"].append(f"暂不支持解析 {path.suffix} 文件: {path.name}")
            payload["profile"]["errors"].append(file_entry["errors"][-1])
            continue

        try:
            workbook = load_workbook(path, data_only=True, read_only=True)
        except Exception as exc:
            file_entry["status"] = "error"
            file_entry["errors"].append(f"无法打开 Excel 文件 {path.name}: {exc}")
            payload["profile"]["errors"].append(file_entry["errors"][-1])
            continue

        file_datasets: list[dict[str, Any]] = []
        for sheet in workbook.worksheets:
            file_entry["sheet_names"].append(sheet.title)
            if getattr(sheet, "max_row", 0) > MAX_ROW_DATASET_CANDIDATES_PER_SHEET:
                warning = (
                    f"{path.name}/{sheet.title} 行数较多；行向 dataset 候选最多保留 "
                    f"{MAX_ROW_DATASET_CANDIDATES_PER_SHEET} 行。"
                )
                file_entry.setdefault("warnings", []).append(warning)
                payload["profile"]["warnings"].append(warning)
            if getattr(sheet, "max_column", 0) > MAX_COLUMN_DATASET_CANDIDATES_PER_SHEET:
                warning = (
                    f"{path.name}/{sheet.title} 列数较多；列向 dataset 候选最多保留 "
                    f"{MAX_COLUMN_DATASET_CANDIDATES_PER_SHEET} 列。"
                )
                file_entry.setdefault("warnings", []).append(warning)
                payload["profile"]["warnings"].append(warning)
            sheet_datasets = _datasets_from_sheet(path, sheet, min_values=min_values)
            file_datasets.extend(sheet_datasets)

        original_file_dataset_count = len(file_datasets)
        if original_file_dataset_count > max_datasets_per_file:
            warning = (
                f"{path.name} 标准化得到 {original_file_dataset_count} 个 dataset；"
                f"为控制上下文和比对成本，仅保留前 {max_datasets_per_file} 个。"
            )
            file_entry.setdefault("warnings", []).append(warning)
            payload["profile"]["warnings"].append(warning)
            file_datasets = file_datasets[:max_datasets_per_file]

        remaining_total = max_total_datasets - len(payload["datasets"])
        if remaining_total <= 0:
            warning = (
                f"已达到全局 dataset 上限 {max_total_datasets}；跳过 {path.name} 的 "
                f"{len(file_datasets)} 个 dataset。"
            )
            file_entry.setdefault("warnings", []).append(warning)
            payload["profile"]["warnings"].append(warning)
            file_datasets = []
        elif len(file_datasets) > remaining_total:
            warning = (
                f"为满足全局 dataset 上限 {max_total_datasets}，{path.name} 仅保留 "
                f"{remaining_total}/{len(file_datasets)} 个 dataset。"
            )
            file_entry.setdefault("warnings", []).append(warning)
            payload["profile"]["warnings"].append(warning)
            file_datasets = file_datasets[:remaining_total]

        file_entry["status"] = "success"
        file_entry["dataset_count"] = len(file_datasets)
        file_entry["original_dataset_count"] = original_file_dataset_count
        payload["datasets"].extend(file_datasets)
        payload["profile"]["loaded_file_count"] += 1
        payload["profile"]["sheet_count"] += len(workbook.worksheets)

    payload["profile"]["dataset_count"] = len(payload["datasets"])
    payload["profile"]["numeric_value_count"] = sum(
        len(dataset.get("values", [])) for dataset in payload["datasets"]
    )
    return payload
